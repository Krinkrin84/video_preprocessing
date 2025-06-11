import os
import argparse
from openpyxl import Workbook
from tqdm import tqdm


def extract_label(filename):
    """Extract the label from filename (the part after last underscore before extension)"""
    base_name = os.path.splitext(filename)[0]  # Remove extension
    parts = base_name.split('_')
    return parts[-1] if len(parts) > 1 else ''


def get_file_data(folder_path):
    """Get all filenames and labels from the specified folder with progress bar"""
    file_data = []
    for entry in tqdm(os.listdir(folder_path), desc="Scanning files"):
        full_path = os.path.join(folder_path, entry)
        if os.path.isfile(full_path):
            label = extract_label(entry)
            file_data.append((entry, label))
    return file_data


def write_to_excel(file_data, output_file):
    """Write filenames and labels to an Excel file with progress bar"""
    wb = Workbook()
    ws = wb.active
    ws.title = "File Data"

    # Write headers
    ws['A1'] = 'filename'
    ws['B1'] = 'label'

    # Write data with progress bar
    for idx, (filename, label) in enumerate(tqdm(file_data, desc="Writing to Excel"), start=2):
        ws.cell(row=idx, column=1, value=filename)
        ws.cell(row=idx, column=2, value=label)

    wb.save(output_file)
    print(f"\nSuccessfully wrote {len(file_data)} records to {output_file}")


def main():
    parser = argparse.ArgumentParser(description='List files in a folder and save to Excel with labels')
    parser.add_argument('folder_path', help='Path to the folder to scan')
    parser.add_argument('--output', default='file_labels.xlsx',
                        help='Output Excel filename (default: file_labels.xlsx)')

    args = parser.parse_args()

    if not os.path.isdir(args.folder_path):
        print(f"Error: {args.folder_path} is not a valid directory")
        return

    print(f"Processing files in: {args.folder_path}")
    file_data = get_file_data(args.folder_path)

    if not file_data:
        print("No files found in the specified folder")
        return

    write_to_excel(file_data, args.output)


if __name__ == '__main__':
    main()